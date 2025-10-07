import pandas as pd
import json
import csv
from io import BytesIO, StringIO
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import zipfile

class DataExporter:
    """Utility class for exporting data in various formats with advanced filtering and metadata"""
    
    def __init__(self):
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.generation_time = datetime.now()
        self.version = "1.0"
    
    def export_to_csv(self, data: List[Dict[str, Any]], filename: str = None, 
                     include_metadata: bool = True, filters_applied: Dict[str, Any] = None) -> str:
        """Export data to CSV format with proper encoding and optional metadata"""
        
        if not data:
            return ""
        
        df = pd.DataFrame(data)
        
        # Create CSV content with UTF-8 BOM for better Excel compatibility
        csv_buffer = StringIO()
        
        # Add metadata header if requested
        if include_metadata:
            csv_buffer.write(f"# Sistema de Reportes Docentes - Exportación CSV\n")
            csv_buffer.write(f"# Fecha de generación: {self.generation_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            csv_buffer.write(f"# Total de registros: {len(data)}\n")
            
            if filters_applied:
                csv_buffer.write(f"# Filtros aplicados: {', '.join(f'{k}={v}' for k, v in filters_applied.items())}\n")
            
            csv_buffer.write(f"# Versión del sistema: {self.version}\n")
            csv_buffer.write("#\n")
        
        # Write CSV data
        df.to_csv(csv_buffer, index=False, encoding='utf-8', lineterminator='\n')
        
        # Return with UTF-8 BOM for Excel compatibility
        csv_content = csv_buffer.getvalue()
        return '\ufeff' + csv_content
    
    def export_to_excel(self, 
                       data: Dict[str, List[Dict[str, Any]]], 
                       filename: str = None,
                       include_charts: bool = True) -> BytesIO:
        """Export data to Excel format with multiple sheets and formatting"""
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # Create summary sheet
            if 'summary' in data:
                summary_df = pd.DataFrame(data['summary'])
                summary_df.to_excel(writer, sheet_name='Resumen', index=False)
                self._format_summary_sheet(writer.book['Resumen'], summary_df)
            
            # Create data sheets
            for sheet_name, sheet_data in data.items():
                if sheet_name == 'summary':
                    continue
                
                if sheet_data:
                    df = pd.DataFrame(sheet_data)
                    safe_sheet_name = self._safe_sheet_name(sheet_name)
                    df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                    
                    # Format the sheet
                    worksheet = writer.book[safe_sheet_name]
                    self._format_data_sheet(worksheet, df)
                    
                    # Add charts if requested
                    if include_charts and len(df) > 1:
                        self._add_chart_to_sheet(worksheet, df)
        
        output.seek(0)
        return output
    
    def export_to_json(self, data: Any, pretty: bool = True) -> str:
        """Export data to JSON format"""
        
        if pretty:
            return json.dumps(data, indent=2, ensure_ascii=False, default=str)
        else:
            return json.dumps(data, ensure_ascii=False, default=str)
    
    def apply_filters(self, forms: List[Any], filters: Dict[str, Any]) -> List[Any]:
        """Apply advanced filters to forms data before export"""
        
        filtered_forms = forms
        
        # Filter by status
        if 'estados' in filters and filters['estados']:
            if 'TODOS' not in filters['estados']:
                filtered_forms = [f for f in filtered_forms if f.estado.value in filters['estados']]
        
        # Filter by date range
        if 'fecha_inicio' in filters and filters['fecha_inicio']:
            fecha_inicio = filters['fecha_inicio']
            if isinstance(fecha_inicio, str):
                fecha_inicio = datetime.fromisoformat(fecha_inicio).date()
            filtered_forms = [f for f in filtered_forms if f.fecha_envio and f.fecha_envio.date() >= fecha_inicio]
        
        if 'fecha_fin' in filters and filters['fecha_fin']:
            fecha_fin = filters['fecha_fin']
            if isinstance(fecha_fin, str):
                fecha_fin = datetime.fromisoformat(fecha_fin).date()
            filtered_forms = [f for f in filtered_forms if f.fecha_envio and f.fecha_envio.date() <= fecha_fin]
        
        # Filter by minimum activity count
        if 'min_actividades' in filters and filters['min_actividades']:
            min_count = int(filters['min_actividades'])
            filtered_forms = [
                f for f in filtered_forms 
                if (len(f.cursos_capacitacion) + len(f.publicaciones) + 
                    len(f.eventos_academicos) + len(f.diseno_curricular) + 
                    len(f.movilidad) + len(f.reconocimientos) + 
                    len(f.certificaciones)) >= min_count
            ]
        
        # Filter by docente name (partial match)
        if 'nombre_docente' in filters and filters['nombre_docente']:
            nombre_filter = filters['nombre_docente'].lower()
            filtered_forms = [
                f for f in filtered_forms 
                if nombre_filter in f.nombre_completo.lower()
            ]
        
        # Filter by email domain
        if 'dominio_email' in filters and filters['dominio_email']:
            dominio = filters['dominio_email'].lower()
            filtered_forms = [
                f for f in filtered_forms 
                if dominio in f.correo_institucional.lower()
            ]
        
        # Filter by specific categories having data
        if 'tiene_cursos' in filters and filters['tiene_cursos']:
            filtered_forms = [f for f in filtered_forms if len(f.cursos_capacitacion) > 0]
        
        if 'tiene_publicaciones' in filters and filters['tiene_publicaciones']:
            filtered_forms = [f for f in filtered_forms if len(f.publicaciones) > 0]
        
        if 'tiene_eventos' in filters and filters['tiene_eventos']:
            filtered_forms = [f for f in filtered_forms if len(f.eventos_academicos) > 0]
        
        return filtered_forms
    
    def export_forms_detailed(self, forms: List[Any], filters_applied: Dict[str, Any] = None) -> Dict[str, List[Dict[str, Any]]]:
        """Export forms with detailed breakdown by category and enhanced metadata"""
        
        export_data = {
            'summary': [],
            'formularios': [],
            'cursos': [],
            'publicaciones': [],
            'eventos': [],
            'disenos_curriculares': [],
            'movilidad': [],
            'reconocimientos': [],
            'certificaciones': []
        }
        
        # Enhanced summary data with metadata
        total_forms = len(forms)
        approved_forms = len([f for f in forms if f.estado.value == 'APROBADO'])
        pending_forms = len([f for f in forms if f.estado.value == 'PENDIENTE'])
        rejected_forms = len([f for f in forms if f.estado.value == 'RECHAZADO'])
        
        # Calculate totals across all categories
        total_cursos = sum(len(f.cursos_capacitacion) for f in forms)
        total_publicaciones = sum(len(f.publicaciones) for f in forms)
        total_eventos = sum(len(f.eventos_academicos) for f in forms)
        total_disenos = sum(len(f.diseno_curricular) for f in forms)
        total_movilidades = sum(len(f.movilidad) for f in forms)
        total_reconocimientos = sum(len(f.reconocimientos) for f in forms)
        total_certificaciones = sum(len(f.certificaciones) for f in forms)
        
        export_data['summary'] = [
            # Metadata section
            {'Métrica': '=== INFORMACIÓN DE EXPORTACIÓN ===', 'Valor': ''},
            {'Métrica': 'Fecha de Generación', 'Valor': self.generation_time.strftime('%Y-%m-%d %H:%M:%S')},
            {'Métrica': 'Versión del Sistema', 'Valor': self.version},
            {'Métrica': 'Timestamp de Exportación', 'Valor': self.timestamp},
            {'Métrica': '', 'Valor': ''},
            
            # Forms summary
            {'Métrica': '=== RESUMEN DE FORMULARIOS ===', 'Valor': ''},
            {'Métrica': 'Total de Formularios', 'Valor': total_forms},
            {'Métrica': 'Formularios Aprobados', 'Valor': approved_forms},
            {'Métrica': 'Formularios Pendientes', 'Valor': pending_forms},
            {'Métrica': 'Formularios Rechazados', 'Valor': rejected_forms},
            {'Métrica': 'Tasa de Aprobación (%)', 'Valor': round((approved_forms / total_forms * 100), 2) if total_forms > 0 else 0},
            {'Métrica': '', 'Valor': ''},
            
            # Activities summary
            {'Métrica': '=== RESUMEN DE ACTIVIDADES ===', 'Valor': ''},
            {'Métrica': 'Total Cursos de Capacitación', 'Valor': total_cursos},
            {'Métrica': 'Total Publicaciones', 'Valor': total_publicaciones},
            {'Métrica': 'Total Eventos Académicos', 'Valor': total_eventos},
            {'Métrica': 'Total Diseños Curriculares', 'Valor': total_disenos},
            {'Métrica': 'Total Experiencias de Movilidad', 'Valor': total_movilidades},
            {'Métrica': 'Total Reconocimientos', 'Valor': total_reconocimientos},
            {'Métrica': 'Total Certificaciones', 'Valor': total_certificaciones},
            {'Métrica': '', 'Valor': ''},
            
            # Derived metrics
            {'Métrica': '=== MÉTRICAS DERIVADAS ===', 'Valor': ''},
            {'Métrica': 'Promedio Cursos por Formulario', 'Valor': round(total_cursos / approved_forms, 2) if approved_forms > 0 else 0},
            {'Métrica': 'Promedio Publicaciones por Formulario', 'Valor': round(total_publicaciones / approved_forms, 2) if approved_forms > 0 else 0},
            {'Métrica': 'Promedio Eventos por Formulario', 'Valor': round(total_eventos / approved_forms, 2) if approved_forms > 0 else 0},
            {'Métrica': 'Total Actividades Registradas', 'Valor': total_cursos + total_publicaciones + total_eventos + total_disenos + total_movilidades + total_reconocimientos + total_certificaciones}
        ]
        
        # Add filter information if provided
        if filters_applied:
            export_data['summary'].extend([
                {'Métrica': '', 'Valor': ''},
                {'Métrica': '=== FILTROS APLICADOS ===', 'Valor': ''}
            ])
            for filter_key, filter_value in filters_applied.items():
                export_data['summary'].append({
                    'Métrica': f'Filtro: {filter_key}', 
                    'Valor': str(filter_value)
                })
        
        # Process each form
        for form in forms:
            # Main form data
            form_data = {
                'ID': form.id,
                'Nombre Completo': form.nombre_completo,
                'Correo Institucional': form.correo_institucional,
                'Estado': form.estado.value,
                'Fecha Envío': form.fecha_envio.strftime('%Y-%m-%d %H:%M') if form.fecha_envio else '',
                'Fecha Revisión': form.fecha_revision.strftime('%Y-%m-%d %H:%M') if form.fecha_revision else '',
                'Revisado Por': form.revisado_por or '',
                'Total Cursos': len(form.cursos_capacitacion),
                'Total Publicaciones': len(form.publicaciones),
                'Total Eventos': len(form.eventos_academicos),
                'Total Diseños': len(form.diseno_curricular),
                'Total Movilidades': len(form.movilidad),
                'Total Reconocimientos': len(form.reconocimientos),
                'Total Certificaciones': len(form.certificaciones)
            }
            export_data['formularios'].append(form_data)
            
            # Detailed data for each category
            for curso in form.cursos_capacitacion:
                export_data['cursos'].append({
                    'Formulario ID': form.id,
                    'Docente': form.nombre_completo,
                    'Nombre del Curso': curso.nombre_curso,
                    'Fecha': curso.fecha.strftime('%Y-%m-%d') if curso.fecha else '',
                    'Horas': curso.horas
                })
            
            for pub in form.publicaciones:
                export_data['publicaciones'].append({
                    'Formulario ID': form.id,
                    'Docente': form.nombre_completo,
                    'Autores': pub.autores,
                    'Título': pub.titulo,
                    'Evento/Revista': pub.evento_revista,
                    'Estatus': pub.estatus.value
                })
            
            for evento in form.eventos_academicos:
                export_data['eventos'].append({
                    'Formulario ID': form.id,
                    'Docente': form.nombre_completo,
                    'Nombre del Evento': evento.nombre_evento,
                    'Fecha': evento.fecha.strftime('%Y-%m-%d') if evento.fecha else '',
                    'Tipo de Participación': evento.tipo_participacion.value
                })
            
            for diseno in form.diseno_curricular:
                export_data['disenos_curriculares'].append({
                    'Formulario ID': form.id,
                    'Docente': form.nombre_completo,
                    'Nombre del Curso': diseno.nombre_curso,
                    'Descripción': diseno.descripcion or ''
                })
            
            for mov in form.movilidad:
                export_data['movilidad'].append({
                    'Formulario ID': form.id,
                    'Docente': form.nombre_completo,
                    'Descripción': mov.descripcion,
                    'Tipo': mov.tipo.value,
                    'Fecha': mov.fecha.strftime('%Y-%m-%d') if mov.fecha else ''
                })
            
            for rec in form.reconocimientos:
                export_data['reconocimientos'].append({
                    'Formulario ID': form.id,
                    'Docente': form.nombre_completo,
                    'Nombre': rec.nombre,
                    'Tipo': rec.tipo.value,
                    'Fecha': rec.fecha.strftime('%Y-%m-%d') if rec.fecha else ''
                })
            
            for cert in form.certificaciones:
                export_data['certificaciones'].append({
                    'Formulario ID': form.id,
                    'Docente': form.nombre_completo,
                    'Nombre': cert.nombre,
                    'Fecha Obtención': cert.fecha_obtencion.strftime('%Y-%m-%d') if cert.fecha_obtencion else '',
                    'Fecha Vencimiento': cert.fecha_vencimiento.strftime('%Y-%m-%d') if cert.fecha_vencimiento else '',
                    'Vigente': 'Sí' if cert.vigente else 'No'
                })
        
        return export_data
    
    def export_metrics_report(self, metrics: Any, period_data: Dict[str, Any] = None) -> Dict[str, Any]:
        """Export comprehensive metrics report"""
        
        report = {
            'fecha_generacion': datetime.now().isoformat(),
            'metricas_generales': {
                'total_formularios': metrics.total_formularios,
                'formularios_pendientes': metrics.formularios_pendientes,
                'formularios_aprobados': metrics.formularios_aprobados,
                'formularios_rechazados': metrics.formularios_rechazados,
                'total_cursos': metrics.total_cursos,
                'total_horas_capacitacion': metrics.total_horas_capacitacion,
                'total_publicaciones': metrics.total_publicaciones,
                'total_eventos': metrics.total_eventos,
                'total_disenos_curriculares': metrics.total_disenos_curriculares,
                'total_movilidades': metrics.total_movilidades,
                'total_reconocimientos': metrics.total_reconocimientos,
                'total_certificaciones': metrics.total_certificaciones
            }
        }
        
        # Add period data if available
        if period_data:
            report['datos_periodo'] = period_data
        
        # Calculate derived metrics
        if metrics.total_formularios > 0:
            report['metricas_derivadas'] = {
                'tasa_aprobacion': round((metrics.formularios_aprobados / metrics.total_formularios) * 100, 2),
                'tasa_rechazo': round((metrics.formularios_rechazados / metrics.total_formularios) * 100, 2),
                'promedio_cursos_por_formulario': round(metrics.total_cursos / metrics.formularios_aprobados, 2) if metrics.formularios_aprobados > 0 else 0,
                'promedio_horas_por_curso': round(metrics.total_horas_capacitacion / metrics.total_cursos, 2) if metrics.total_cursos > 0 else 0,
                'promedio_publicaciones_por_formulario': round(metrics.total_publicaciones / metrics.formularios_aprobados, 2) if metrics.formularios_aprobados > 0 else 0
            }
        
        return report
    
    def export_with_advanced_options(self, 
                                   forms: List[Any], 
                                   export_format: str = 'excel',
                                   filters: Dict[str, Any] = None,
                                   include_metadata: bool = True,
                                   include_charts: bool = True,
                                   custom_filename: str = None) -> BytesIO:
        """
        Advanced export method with comprehensive filtering and options
        
        Args:
            forms: List of form objects to export
            export_format: 'csv', 'excel', 'json', or 'package'
            filters: Dictionary of filters to apply
            include_metadata: Whether to include metadata in export
            include_charts: Whether to include charts in Excel export
            custom_filename: Custom filename prefix
            
        Returns:
            BytesIO object with the exported data
        """
        
        # Apply filters if provided
        if filters:
            filtered_forms = self.apply_filters(forms, filters)
        else:
            filtered_forms = forms
            filters = {}
        
        # Generate filename
        if custom_filename:
            base_filename = f"{custom_filename}_{self.timestamp}"
        else:
            base_filename = f"exportacion_docentes_{self.timestamp}"
        
        output = BytesIO()
        
        if export_format.lower() == 'csv':
            # Enhanced CSV export
            basic_data = []
            for form in filtered_forms:
                basic_data.append({
                    'ID': form.id,
                    'Nombre_Completo': form.nombre_completo,
                    'Correo_Institucional': form.correo_institucional,
                    'Estado': form.estado.value,
                    'Fecha_Envio': form.fecha_envio.strftime('%Y-%m-%d %H:%M:%S') if form.fecha_envio else '',
                    'Fecha_Revision': form.fecha_revision.strftime('%Y-%m-%d %H:%M:%S') if form.fecha_revision else '',
                    'Revisado_Por': form.revisado_por or '',
                    'Total_Cursos': len(form.cursos_capacitacion),
                    'Total_Publicaciones': len(form.publicaciones),
                    'Total_Eventos': len(form.eventos_academicos),
                    'Total_Disenos_Curriculares': len(form.diseno_curricular),
                    'Total_Movilidades': len(form.movilidad),
                    'Total_Reconocimientos': len(form.reconocimientos),
                    'Total_Certificaciones': len(form.certificaciones),
                    'Total_Actividades': (
                        len(form.cursos_capacitacion) + len(form.publicaciones) +
                        len(form.eventos_academicos) + len(form.diseno_curricular) +
                        len(form.movilidad) + len(form.reconocimientos) +
                        len(form.certificaciones)
                    )
                })
            
            csv_content = self.export_to_csv(basic_data, include_metadata=include_metadata, filters_applied=filters)
            output.write(csv_content.encode('utf-8'))
        
        elif export_format.lower() == 'excel':
            # Enhanced Excel export
            detailed_data = self.export_forms_detailed(filtered_forms, filters_applied=filters)
            excel_output = self.export_to_excel(detailed_data, include_charts=include_charts)
            output.write(excel_output.getvalue())
        
        elif export_format.lower() == 'json':
            # Enhanced JSON export
            json_data = {
                'metadata': {
                    'fecha_exportacion': self.generation_time.isoformat(),
                    'timestamp': self.timestamp,
                    'version_sistema': self.version,
                    'total_registros_originales': len(forms),
                    'total_registros_filtrados': len(filtered_forms),
                    'filtros_aplicados': filters,
                    'generado_por': 'Sistema de Reportes Docentes'
                } if include_metadata else None,
                'datos': []
            }
            
            for form in filtered_forms:
                form_data = {
                    'id': form.id,
                    'nombre_completo': form.nombre_completo,
                    'correo_institucional': form.correo_institucional,
                    'estado': form.estado.value,
                    'fecha_envio': form.fecha_envio.isoformat() if form.fecha_envio else None,
                    'fecha_revision': form.fecha_revision.isoformat() if form.fecha_revision else None,
                    'revisado_por': form.revisado_por,
                    'actividades': {
                        'cursos_capacitacion': [
                            {
                                'nombre_curso': c.nombre_curso,
                                'fecha': c.fecha.isoformat() if c.fecha else None,
                                'horas': c.horas
                            } for c in form.cursos_capacitacion
                        ],
                        'publicaciones': [
                            {
                                'autores': p.autores,
                                'titulo': p.titulo,
                                'evento_revista': p.evento_revista,
                                'estatus': p.estatus.value
                            } for p in form.publicaciones
                        ],
                        'eventos_academicos': [
                            {
                                'nombre_evento': e.nombre_evento,
                                'fecha': e.fecha.isoformat() if e.fecha else None,
                                'tipo_participacion': e.tipo_participacion.value
                            } for e in form.eventos_academicos
                        ],
                        'diseno_curricular': [
                            {
                                'nombre_curso': d.nombre_curso,
                                'descripcion': d.descripcion
                            } for d in form.diseno_curricular
                        ],
                        'movilidad': [
                            {
                                'descripcion': m.descripcion,
                                'tipo': m.tipo.value,
                                'fecha': m.fecha.isoformat() if m.fecha else None
                            } for m in form.movilidad
                        ],
                        'reconocimientos': [
                            {
                                'nombre': r.nombre,
                                'tipo': r.tipo.value,
                                'fecha': r.fecha.isoformat() if r.fecha else None
                            } for r in form.reconocimientos
                        ],
                        'certificaciones': [
                            {
                                'nombre': c.nombre,
                                'fecha_obtencion': c.fecha_obtencion.isoformat() if c.fecha_obtencion else None,
                                'fecha_vencimiento': c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else None,
                                'vigente': c.vigente
                            } for c in form.certificaciones
                        ]
                    }
                }
                json_data['datos'].append(form_data)
            
            json_content = self.export_to_json(json_data, pretty=True)
            output.write(json_content.encode('utf-8'))
        
        elif export_format.lower() == 'package':
            # Create comprehensive package with all formats
            return self.create_comprehensive_package(filtered_forms, filters, include_metadata, include_charts, base_filename)
        
        output.seek(0)
        return output
    
    def create_comprehensive_package(self, 
                                   forms: List[Any], 
                                   filters: Dict[str, Any],
                                   include_metadata: bool,
                                   include_charts: bool,
                                   base_filename: str) -> BytesIO:
        """Create a comprehensive export package with all formats and documentation"""
        
        package = BytesIO()
        
        with zipfile.ZipFile(package, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            
            # 1. Basic CSV export
            basic_data = []
            for form in forms:
                basic_data.append({
                    'ID': form.id,
                    'Nombre_Completo': form.nombre_completo,
                    'Correo_Institucional': form.correo_institucional,
                    'Estado': form.estado.value,
                    'Fecha_Envio': form.fecha_envio.strftime('%Y-%m-%d %H:%M:%S') if form.fecha_envio else '',
                    'Total_Actividades': (
                        len(form.cursos_capacitacion) + len(form.publicaciones) +
                        len(form.eventos_academicos) + len(form.diseno_curricular) +
                        len(form.movilidad) + len(form.reconocimientos) +
                        len(form.certificaciones)
                    )
                })
            
            csv_content = self.export_to_csv(basic_data, include_metadata=include_metadata, filters_applied=filters)
            zip_file.writestr(f'{base_filename}_basico.csv', csv_content.encode('utf-8'))
            
            # 2. Detailed Excel export
            detailed_data = self.export_forms_detailed(forms, filters_applied=filters)
            excel_content = self.export_to_excel(detailed_data, include_charts=include_charts)
            zip_file.writestr(f'{base_filename}_detallado.xlsx', excel_content.getvalue())
            
            # 3. Complete JSON export
            json_export = self.export_with_advanced_options(
                forms, 'json', filters, include_metadata, include_charts, base_filename
            )
            zip_file.writestr(f'{base_filename}_completo.json', json_export.getvalue())
            
            # 4. Individual category CSV files
            categories = {
                'cursos': [],
                'publicaciones': [],
                'eventos': [],
                'disenos_curriculares': [],
                'movilidad': [],
                'reconocimientos': [],
                'certificaciones': []
            }
            
            for form in forms:
                # Extract data for each category
                for curso in form.cursos_capacitacion:
                    categories['cursos'].append({
                        'Formulario_ID': form.id,
                        'Docente': form.nombre_completo,
                        'Nombre_Curso': curso.nombre_curso,
                        'Fecha': curso.fecha.strftime('%Y-%m-%d') if curso.fecha else '',
                        'Horas': curso.horas
                    })
                
                for pub in form.publicaciones:
                    categories['publicaciones'].append({
                        'Formulario_ID': form.id,
                        'Docente': form.nombre_completo,
                        'Autores': pub.autores,
                        'Titulo': pub.titulo,
                        'Evento_Revista': pub.evento_revista,
                        'Estatus': pub.estatus.value
                    })
                
                # Continue for other categories...
                for evento in form.eventos_academicos:
                    categories['eventos'].append({
                        'Formulario_ID': form.id,
                        'Docente': form.nombre_completo,
                        'Nombre_Evento': evento.nombre_evento,
                        'Fecha': evento.fecha.strftime('%Y-%m-%d') if evento.fecha else '',
                        'Tipo_Participacion': evento.tipo_participacion.value
                    })
            
            # Save individual category files
            for category, data in categories.items():
                if data:  # Only create file if there's data
                    category_csv = self.export_to_csv(data, include_metadata=False)
                    zip_file.writestr(f'{base_filename}_{category}.csv', category_csv.encode('utf-8'))
            
            # 5. Enhanced README with export details
            readme_content = self._create_enhanced_readme(forms, filters, base_filename)
            zip_file.writestr('README.txt', readme_content)
            
            # 6. Export summary report
            summary_report = self._create_export_summary(forms, filters)
            zip_file.writestr('RESUMEN_EXPORTACION.txt', summary_report)
        
        package.seek(0)
        return package
    
    def create_export_package(self, 
                            forms: List[Any], 
                            metrics: Any,
                            include_detailed: bool = True,
                            include_charts: bool = True) -> BytesIO:
        """Create a complete export package with multiple formats"""
        
        package = BytesIO()
        
        with zipfile.ZipFile(package, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            
            # Export basic CSV
            basic_data = []
            for form in forms:
                basic_data.append({
                    'ID': form.id,
                    'Nombre': form.nombre_completo,
                    'Email': form.correo_institucional,
                    'Estado': form.estado.value,
                    'Fecha_Envio': form.fecha_envio.strftime('%Y-%m-%d') if form.fecha_envio else '',
                    'Total_Actividades': (
                        len(form.cursos_capacitacion) + len(form.publicaciones) +
                        len(form.eventos_academicos) + len(form.diseno_curricular) +
                        len(form.movilidad) + len(form.reconocimientos) +
                        len(form.certificaciones)
                    )
                })
            
            csv_content = self.export_to_csv(basic_data)
            zip_file.writestr(f'formularios_basico_{self.timestamp}.csv', csv_content)
            
            # Export detailed Excel if requested
            if include_detailed:
                detailed_data = self.export_forms_detailed(forms)
                excel_content = self.export_to_excel(detailed_data, include_charts=include_charts)
                zip_file.writestr(f'formularios_detallado_{self.timestamp}.xlsx', excel_content.getvalue())
            
            # Export metrics report
            metrics_report = self.export_metrics_report(metrics)
            json_content = self.export_to_json(metrics_report)
            zip_file.writestr(f'reporte_metricas_{self.timestamp}.json', json_content)
            
            # Create README
            readme_content = self._create_readme()
            zip_file.writestr('README.txt', readme_content)
        
        package.seek(0)
        return package
    
    def _format_summary_sheet(self, worksheet, df):
        """Format the summary sheet in Excel"""
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_data_sheet(self, worksheet, df):
        """Format data sheets in Excel"""
        
        # Header formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_chart_to_sheet(self, worksheet, df):
        """Add a simple chart to the Excel sheet"""
        
        try:
            # Only add chart if we have numeric data
            numeric_columns = df.select_dtypes(include=['int64', 'float64']).columns
            
            if len(numeric_columns) > 0:
                chart = BarChart()
                chart.title = "Distribución de Datos"
                chart.style = 10
                chart.x_axis.title = 'Categorías'
                chart.y_axis.title = 'Valores'
                
                # Use first numeric column for chart
                data_col = numeric_columns[0]
                data = Reference(worksheet, min_col=df.columns.get_loc(data_col) + 1, 
                               min_row=1, max_row=min(len(df) + 1, 20))
                chart.add_data(data, titles_from_data=True)
                
                # Position chart to the right of data
                chart_position = f"{chr(ord('A') + len(df.columns) + 2)}2"
                worksheet.add_chart(chart, chart_position)
        except Exception:
            # If chart creation fails, continue without chart
            pass
    
    def _safe_sheet_name(self, name: str) -> str:
        """Create a safe sheet name for Excel"""
        
        # Remove invalid characters and limit length
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        safe_name = name
        
        for char in invalid_chars:
            safe_name = safe_name.replace(char, '_')
        
        return safe_name[:31]  # Excel sheet name limit
    
    def _create_readme(self) -> str:
        """Create README content for export package"""
        
        return f"""
SISTEMA DE REPORTES DOCENTES - EXPORTACIÓN DE DATOS
==================================================

Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

CONTENIDO DEL PAQUETE:
---------------------

1. formularios_basico_{self.timestamp}.csv
   - Datos básicos de todos los formularios
   - Formato: CSV (separado por comas)
   - Incluye: ID, nombre, email, estado, fecha de envío, total de actividades

2. formularios_detallado_{self.timestamp}.xlsx
   - Datos completos con múltiples hojas
   - Formato: Excel con formato y gráficos
   - Hojas incluidas:
     * Resumen: Métricas generales
     * Formularios: Datos principales de cada formulario
     * Cursos: Detalle de cursos de capacitación
     * Publicaciones: Detalle de publicaciones
     * Eventos: Detalle de eventos académicos
     * Diseños Curriculares: Detalle de diseños curriculares
     * Movilidad: Detalle de experiencias de movilidad
     * Reconocimientos: Detalle de reconocimientos
     * Certificaciones: Detalle de certificaciones

3. reporte_metricas_{self.timestamp}.json
   - Reporte completo de métricas y estadísticas
   - Formato: JSON estructurado
   - Incluye métricas generales y derivadas

NOTAS:
------
- Todos los archivos están codificados en UTF-8
- Las fechas están en formato YYYY-MM-DD
- Los datos corresponden únicamente a formularios con el estado seleccionado
- Para soporte técnico, contacte al administrador del sistema

Sistema de Reportes Docentes v1.0
"""
    
    def _create_enhanced_readme(self, forms: List[Any], filters: Dict[str, Any], base_filename: str) -> str:
        """Create enhanced README with detailed export information"""
        
        total_forms = len(forms)
        status_counts = {}
        for form in forms:
            status = form.estado.value
            status_counts[status] = status_counts.get(status, 0) + 1
        
        return f"""
SISTEMA DE REPORTES DOCENTES - EXPORTACIÓN AVANZADA
==================================================

INFORMACIÓN DE EXPORTACIÓN:
--------------------------
Fecha de generación: {self.generation_time.strftime('%Y-%m-%d %H:%M:%S')}
Timestamp: {self.timestamp}
Versión del sistema: {self.version}
Nombre base de archivos: {base_filename}

DATOS EXPORTADOS:
----------------
Total de formularios: {total_forms}
Distribución por estado:
{chr(10).join(f'  - {status}: {count}' for status, count in status_counts.items())}

FILTROS APLICADOS:
-----------------
{chr(10).join(f'{key}: {value}' for key, value in filters.items()) if filters else 'Ningún filtro aplicado'}

CONTENIDO DEL PAQUETE:
---------------------

ARCHIVOS PRINCIPALES:
1. {base_filename}_basico.csv
   - Datos básicos de formularios con metadatos
   - Formato: CSV con BOM UTF-8 para compatibilidad con Excel
   - Incluye comentarios con información de exportación

2. {base_filename}_detallado.xlsx
   - Archivo Excel completo con múltiples hojas
   - Incluye gráficos y formato profesional
   - Hoja de resumen con métricas completas

3. {base_filename}_completo.json
   - Exportación completa en formato JSON
   - Incluye metadatos y estructura jerárquica
   - Todas las actividades por formulario

ARCHIVOS POR CATEGORÍA:
4. {base_filename}_cursos.csv - Cursos de capacitación
5. {base_filename}_publicaciones.csv - Publicaciones académicas
6. {base_filename}_eventos.csv - Eventos académicos
7. {base_filename}_disenos_curriculares.csv - Diseños curriculares
8. {base_filename}_movilidad.csv - Experiencias de movilidad
9. {base_filename}_reconocimientos.csv - Reconocimientos
10. {base_filename}_certificaciones.csv - Certificaciones

DOCUMENTACIÓN:
11. README.txt - Este archivo
12. RESUMEN_EXPORTACION.txt - Resumen estadístico de la exportación

CARACTERÍSTICAS TÉCNICAS:
------------------------
- Codificación: UTF-8 con BOM para compatibilidad
- Formato de fechas: ISO 8601 (YYYY-MM-DD HH:MM:SS)
- Separador CSV: Coma (,)
- Compresión: ZIP con deflate
- Metadatos incluidos en todos los formatos

INSTRUCCIONES DE USO:
--------------------
1. Extraiga todos los archivos del ZIP
2. Use el archivo Excel para análisis visual
3. Use los CSV individuales para análisis específicos
4. Use el JSON para integración con otros sistemas
5. Consulte el resumen de exportación para estadísticas

SOPORTE TÉCNICO:
---------------
Para soporte técnico o preguntas sobre los datos exportados,
contacte al administrador del Sistema de Reportes Docentes.

Sistema de Reportes Docentes v{self.version}
Exportación generada automáticamente
"""
    
    def _create_export_summary(self, forms: List[Any], filters: Dict[str, Any]) -> str:
        """Create detailed export summary report"""
        
        total_forms = len(forms)
        status_counts = {}
        activity_totals = {
            'cursos': 0,
            'publicaciones': 0,
            'eventos': 0,
            'disenos': 0,
            'movilidad': 0,
            'reconocimientos': 0,
            'certificaciones': 0
        }
        
        for form in forms:
            status = form.estado.value
            status_counts[status] = status_counts.get(status, 0) + 1
            
            activity_totals['cursos'] += len(form.cursos_capacitacion)
            activity_totals['publicaciones'] += len(form.publicaciones)
            activity_totals['eventos'] += len(form.eventos_academicos)
            activity_totals['disenos'] += len(form.diseno_curricular)
            activity_totals['movilidad'] += len(form.movilidad)
            activity_totals['reconocimientos'] += len(form.reconocimientos)
            activity_totals['certificaciones'] += len(form.certificaciones)
        
        total_activities = sum(activity_totals.values())
        
        return f"""
RESUMEN ESTADÍSTICO DE EXPORTACIÓN
=================================

INFORMACIÓN GENERAL:
-------------------
Fecha de exportación: {self.generation_time.strftime('%Y-%m-%d %H:%M:%S')}
Timestamp: {self.timestamp}
Total de formularios exportados: {total_forms}

DISTRIBUCIÓN POR ESTADO:
-----------------------
{chr(10).join(f'{status}: {count} ({count/total_forms*100:.1f}%)' for status, count in status_counts.items())}

ACTIVIDADES REGISTRADAS:
-----------------------
Cursos de Capacitación: {activity_totals['cursos']}
Publicaciones: {activity_totals['publicaciones']}
Eventos Académicos: {activity_totals['eventos']}
Diseños Curriculares: {activity_totals['disenos']}
Experiencias de Movilidad: {activity_totals['movilidad']}
Reconocimientos: {activity_totals['reconocimientos']}
Certificaciones: {activity_totals['certificaciones']}

TOTAL DE ACTIVIDADES: {total_activities}

MÉTRICAS DERIVADAS:
------------------
Promedio de actividades por formulario: {total_activities/total_forms:.2f}
Formulario con más actividades: {max((len(f.cursos_capacitacion) + len(f.publicaciones) + len(f.eventos_academicos) + len(f.diseno_curricular) + len(f.movilidad) + len(f.reconocimientos) + len(f.certificaciones)) for f in forms) if forms else 0}
Formulario con menos actividades: {min((len(f.cursos_capacitacion) + len(f.publicaciones) + len(f.eventos_academicos) + len(f.diseno_curricular) + len(f.movilidad) + len(f.reconocimientos) + len(f.certificaciones)) for f in forms) if forms else 0}

FILTROS APLICADOS:
-----------------
{chr(10).join(f'{key}: {value}' for key, value in filters.items()) if filters else 'No se aplicaron filtros'}

CALIDAD DE DATOS:
----------------
Formularios con email válido: {len([f for f in forms if '@' in f.correo_institucional])}
Formularios con fecha de envío: {len([f for f in forms if f.fecha_envio])}
Formularios sin actividades: {len([f for f in forms if (len(f.cursos_capacitacion) + len(f.publicaciones) + len(f.eventos_academicos) + len(f.diseno_curricular) + len(f.movilidad) + len(f.reconocimientos) + len(f.certificaciones)) == 0])}

ARCHIVOS GENERADOS:
------------------
- 1 archivo CSV básico
- 1 archivo Excel detallado
- 1 archivo JSON completo
- {len([k for k, v in activity_totals.items() if v > 0])} archivos CSV por categoría
- 2 archivos de documentación

Total de archivos en el paquete: {4 + len([k for k, v in activity_totals.items() if v > 0])}

---
Exportación completada exitosamente
Sistema de Reportes Docentes v{self.version}
"""